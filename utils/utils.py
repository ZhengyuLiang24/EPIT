import numpy as np
import torch
from pathlib import Path
import logging
from option import args
from einops import rearrange
import openpyxl
import torch.nn.functional as F
from skimage.metrics import peak_signal_noise_ratio as cal_PSNR
from skimage.metrics import structural_similarity as cal_SSIM


class ExcelFile(object):
    def __init__(self):
        self.xlsx_file = openpyxl.Workbook()
        self.worksheet = self.xlsx_file.active
        self.worksheet.title = 'sheet1'
        self.header_list = ['Datasets', 'Scenes', 'PSNR', 'SSIM', 'LPIPS']

        self.sum = 1
        self.worksheet.cell(self.sum, 1, 'Datasets')
        self.worksheet.cell(self.sum, 2, 'Scenes')
        self.worksheet.column_dimensions['A'].width = 16
        self.worksheet.column_dimensions['B'].width = 22
        self.add_count(1)

    def write_sheet(self, test_name, LF_name, metric_name, metric_score):
        self.worksheet.cell(self.sum, 1, test_name)
        self.worksheet.cell(self.sum, 2, LF_name)

        # self.worksheet.col(self.header_list.index(metric_name)).width = 256 * 10
        self.worksheet.cell(1, self.header_list.index(metric_name)+1, metric_name)
        self.worksheet.cell(self.sum, self.header_list.index(metric_name)+1, '%.6f' % metric_score)

    def add_count(self, num):
        self.sum = self.sum + num


def get_logger(log_dir, args):
    '''LOG '''
    logger = logging.getLogger(args.model_name)
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler = logging.FileHandler('%s/%s.txt' % (log_dir, args.model_name))
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def create_dir(args):
    log_dir = Path(args.path_log)
    log_dir.mkdir(exist_ok=True)
    if args.task == 'SSR':
        task_path = 'SSR_' + str(args.angRes_in) + 'x' + str(args.angRes_in) + '_' + str(args.scale_factor) + 'x'
    elif args.task == 'ASR':
        task_path = 'ASR_' + str(args.angRes_in) + 'x' + str(args.angRes_in) + '_' + str(args.angRes_out) + 'x' + str(args.angRes_out)
    
    log_dir = log_dir.joinpath(task_path)
    log_dir.mkdir(exist_ok=True)
    log_dir = log_dir.joinpath(args.model_name)
    log_dir.mkdir(exist_ok=True)

    checkpoints_dir = log_dir.joinpath('checkpoints/')
    checkpoints_dir.mkdir(exist_ok=True)

    results_dir = log_dir.joinpath('results/')
    results_dir.mkdir(exist_ok=True)

    return log_dir, checkpoints_dir, results_dir


class Logger():
    def __init__(self, log_dir, args):
        self.logger = get_logger(log_dir, args)

    def log_string(self, str):
        if args.local_rank <= 0:
            self.logger.info(str)
            print(str)


def cal_metrics(args, label, out):
    [B, _, U, V, H, W] = label.size()
    try:
        label = LF_rgb2ycbcr(label)
        out = LF_rgb2ycbcr(out)
    except:
        pass
    label = label[:, 0, :, :, :, :].data.cpu()
    out = out[:, 0, :, :, :, :].data.cpu()

    PSNR = np.zeros(shape=(B, U, V), dtype='float32')
    SSIM = np.zeros(shape=(B, U, V), dtype='float32')
    for b in range(B):
        for u in range(U):
            for v in range(V):
                PSNR[b, u, v] = cal_PSNR(label[b, u, v, :, :].numpy(), out[b, u, v, :, :].numpy())
                if args.task == 'SSR':
                    SSIM[b, u, v] = cal_SSIM(label[b, u, v, :, :].numpy(), out[b, u, v, :, :].numpy(),
                                             gaussian_weights=True)
                elif args.task == 'ASR':
                    SSIM[b, u, v] = cal_SSIM(label[b, u, v, :, :].numpy(), out[b, u, v, :, :].numpy(),
                                             gaussian_weights=True, sigma=1.5, use_sample_covariance=False)

    if args.task=='ASR':
        for u in range(0, args.angRes_out, (args.angRes_out - 1) // (args.angRes_in - 1)):
            for v in range(0, args.angRes_out, (args.angRes_out - 1) // (args.angRes_in - 1)):
                PSNR[:, u, v] = 0
                SSIM[:, u, v] = 0

    PSNR_mean = PSNR.sum() / np.sum(PSNR > 0)
    SSIM_mean = SSIM.sum() / np.sum(SSIM > 0)

    return PSNR_mean, SSIM_mean


class LF_divide_integrate(object):
    def __init__(self, scale, patch_size, stride):
        self.scale = scale
        self.patch_size = patch_size
        self.stride = stride
        self.bdr = (patch_size - stride) // 2
        self.pad = torch.nn.ReflectionPad2d(padding=(self.bdr, self.bdr + stride - 1, self.bdr, self.bdr + stride - 1))

    def LFdivide(self, LF):
        assert LF.size(0) == 1, 'The batch_size of LF for test requires to be one!'
        LF = LF.squeeze(0)
        [c, u, v, h, w] = LF.size()

        LF = rearrange(LF, 'c u v h w -> (c u v) 1 h w')
        self.numU = (h + self.bdr * 2 - 1) // self.stride
        self.numV = (w + self.bdr * 2 - 1) // self.stride

        # LF_pad = self.pad(LF)
        LF_pad = ImageExtend(LF, [self.bdr, self.bdr + self.stride - 1, self.bdr, self.bdr + self.stride - 1])
        LF_divided = F.unfold(LF_pad, kernel_size=self.patch_size, stride=self.stride)
        LF_divided = rearrange(LF_divided, '(c u v) (h w) (numU numV) -> (numU numV) c u v h w', u=u, v=v,
                               h=self.patch_size, w=self.patch_size, numU=self.numU, numV=self.numV)
        return LF_divided

    def LFintegrate(self, LF_divided):
        LF_divided = LF_divided[:, :, :, :, self.bdr*self.scale:(self.bdr+self.stride)*self.scale,
                                self.bdr*self.scale:(self.bdr+self.stride)*self.scale]
        LF = rearrange(LF_divided, '(numU numV) c u v h w -> c u v (numU h) (numV w)',
                       numU=self.numU, numV=self.numV)
        return LF


def ImageExtend(Im, bdr):
    [_, _, h, w] = Im.size()
    Im_lr = torch.flip(Im, dims=[-1])
    Im_ud = torch.flip(Im, dims=[-2])
    Im_diag = torch.flip(Im, dims=[-1, -2])

    Im_up = torch.cat((Im_diag, Im_ud, Im_diag), dim=-1)
    Im_mid = torch.cat((Im_lr, Im, Im_lr), dim=-1)
    Im_down = torch.cat((Im_diag, Im_ud, Im_diag), dim=-1)
    Im_Ext = torch.cat((Im_up, Im_mid, Im_down), dim=-2)
    Im_out = Im_Ext[:, :, h - bdr[0]: 2 * h + bdr[1], w - bdr[2]: 2 * w + bdr[3]]
    return Im_out


def LF_rgb2ycbcr(x):
    y = torch.zeros_like(x)
    y[:,0,:,:,:,:] =  65.481 * x[:,0,:,:,:,:] + 128.553 * x[:,1,:,:,:,:] +  24.966 * x[:,2,:,:,:,:] +  16.0
    y[:,1,:,:,:,:] = -37.797 * x[:,0,:,:,:,:] -  74.203 * x[:,1,:,:,:,:] + 112.000 * x[:,2,:,:,:,:] + 128.0
    y[:,2,:,:,:,:] = 112.000 * x[:,0,:,:,:,:] -  93.786 * x[:,1,:,:,:,:] -  18.214 * x[:,2,:,:,:,:] + 128.0

    y = y / 255.0
    return y


def LF_ycbcr2rgb(x):
    mat = np.array(
        [[65.481, 128.553, 24.966],
         [-37.797, -74.203, 112.0],
         [112.0, -93.786, -18.214]])
    mat_inv = np.linalg.inv(mat)
    offset = np.matmul(mat_inv, np.array([16, 128, 128]))
    mat_inv = mat_inv * 255

    y = torch.zeros_like(x)
    y[:,0,:,:,:,:] = mat_inv[0,0] * x[:,0,:,:,:,:] + mat_inv[0,1] * x[:,1,:,:,:,:] + mat_inv[0,2] * x[:,2,:,:,:,:] - offset[0]
    y[:,1,:,:,:,:] = mat_inv[1,0] * x[:,0,:,:,:,:] + mat_inv[1,1] * x[:,1,:,:,:,:] + mat_inv[1,2] * x[:,2,:,:,:,:] - offset[1]
    y[:,2,:,:,:,:] = mat_inv[2,0] * x[:,0,:,:,:,:] + mat_inv[2,1] * x[:,1,:,:,:,:] + mat_inv[2,2] * x[:,2,:,:,:,:] - offset[2]
    return y


def LF_interpolate(LF, scale_factor, mode):
    [b, c, u, v, h, w] = LF.size()
    LF = rearrange(LF, 'b c u v h w -> (b u v) c h w')
    LF_upscale = F.interpolate(LF, scale_factor=scale_factor, mode=mode, align_corners=False)
    LF_upscale = rearrange(LF_upscale, '(b u v) c h w -> b c u v h w', u=u, v=v)
    return LF_upscale